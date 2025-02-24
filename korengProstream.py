import streamlit as st
import openpyxl
import edge_tts
import asyncio
import os
from pathlib import Path
import base64
import pygame.mixer
import time
from gtts import gTTS  # 파일 상단에 추가
import subprocess

class KorEngPlayer:
    def __init__(self):
        # 다크 테마 설정
        st.set_page_config(
            page_title="English Learning",
            layout="wide",
            initial_sidebar_state="collapsed"
        )
        
        # CSS 스타일 적용
        st.markdown("""
            <style>
                .stApp {
                    background-color: #0E1117;
                    color: #FFFFFF;
                }
                .main .block-container {
                    max-width: 720px;
                    padding-top: 2rem;
                    padding-right: 1rem;
                    padding-left: 1rem;
                    padding-bottom: 2rem;
                }
                .st-emotion-cache-q8sbsg p {
                    color: #FFFFFF !important;
                }
                .st-emotion-cache-183lzff {
                    color: #FFFFFF;
                }
                .stMarkdown {
                    color: #FFFFFF;
                }
                .stSelectbox [data-baseweb="select"] {
                    max-width: 720px;
                }
                .stSlider {
                    max-width: 720px;
                }
                .stButton > button {
                    width: 150px !important;
                    display: inline-flex;
                    justify-content: center;
                    align-items: center;
                    color: #FFFFFF;
                    background-color: #1E1E1E;
                    border: 1px solid #4A4A4A;
                }
                .stButton > button:hover {
                    background-color: #2E2E2E;
                    border-color: #6A6A6A;
                }
                div[data-testid="column"] {
                    display: flex;
                    justify-content: center;
                    padding: 0 5px !important;
                }
                .title {
                    font-size: 20px !important;
                }
                .metric-label {
                    font-size: 14px !important;
                }
                .metric-value {
                    font-size: 16px !important;
                }
                /* 셀렉트박스와 체크박스 레이블 색상 */
                .stSelectbox label, .stCheckbox label {
                    color: #FFFFFF !important;
                }
            </style>
        """, unsafe_allow_html=True)
        
        # 중앙 정렬을 위한 컨테이너
        container = st.container()
        with container:
            st.markdown('<h2 class="title">English Learning</h2>', unsafe_allow_html=True)
        
        # pygame 초기화 수정 - 오디오 없이도 실행되도록
        try:
            os.environ['SDL_AUDIODRIVER'] = 'dummy'
            pygame.mixer.init()
        except Exception as e:
            st.warning("오디오 장치를 찾을 수 없습니다. 음성이 재생되지 않을 수 있습니다.")
        
        # 세션 상태 초기화
        if 'playing' not in st.session_state:
            st.session_state.playing = False
        if 'current_row' not in st.session_state:
            st.session_state.current_row = 2
        if 'repeat' not in st.session_state:
            st.session_state.repeat = False
        if 'last_played_row' not in st.session_state:
            st.session_state.last_played_row = None
            
        # 임시 디렉토리 설정
        self.temp_dir = Path("temp_audio")
        self.temp_dir.mkdir(exist_ok=True)
        
        # 엑셀 파일 로드 - 상대 경로 사용
        try:
            excel_path = 'korengpro.xlsx'  # 현재 작업 디렉토리 기준
            if not os.path.exists(excel_path):
                st.error(f"엑셀 파일을 찾을 수 없습니다. 현재 디렉토리: {os.getcwd()}")
                st.error("'korengpro.xlsx' 파일을 업로드해주세요.")
                uploaded_file = st.file_uploader("Excel 파일 선택", type=['xlsx'])
                if uploaded_file:
                    with open('korengpro.xlsx', 'wb') as f:
                        f.write(uploaded_file.getvalue())
                    excel_path = 'korengpro.xlsx'
                else:
                    st.stop()
            
            self.workbook = openpyxl.load_workbook(excel_path)
            self.sheet_names = self.workbook.sheetnames[:10]
        except Exception as e:
            st.error(f"엑셀 파일 로드 실패: {e}")
            st.stop()
        
        self.setup_interface()

    def cleanup_file(self, filename):
        try:
            if os.path.exists(filename):
                for _ in range(3):
                    try:
                        os.remove(filename)
                        break
                    except:
                        time.sleep(0.1)
        except Exception as e:
            st.error(f"파일 삭제 실패: {filename}")

    def play_audio(self, filename):
        try:
            if os.path.exists(filename):
                with open(filename, 'rb') as f:
                    audio_bytes = f.read()
                self.audio_container.audio(audio_bytes, format='audio/mp3', start_time=0)
        except Exception as e:
            st.error(f"음성 재생 오류: {e}")

    async def generate_speech(self, text, language):
        try:
            voice = "ko-KR-SunHiNeural" if language == "ko" else "en-US-JennyNeural"
            filename = self.temp_dir / f"temp_{language}_{hash(text)}.mp3"
            
            communicate = edge_tts.Communicate(text, voice)
            await communicate.save(str(filename))
            await asyncio.sleep(0.1)
            
            return str(filename)
        except Exception as e:
            st.error(f"음성 생성 오류: {e}")
            return None

    def setup_interface(self):
        # 이전 선택된 시트 추적
        if 'previous_sheet' not in st.session_state:
            st.session_state.previous_sheet = None
        
        # 주제 선택
        selected_sheet = st.selectbox(
            "주제를 선택하세요",
            ["선택하세요"] + self.sheet_names,
            key='sheet_selector'
        )
        
        # 시트가 변경되었을 때 현재 행 초기화
        if selected_sheet != st.session_state.previous_sheet:
            st.session_state.current_row = 2
            st.session_state.previous_sheet = selected_sheet
        
        if selected_sheet and selected_sheet != "선택하세요":
            sheet = self.workbook[selected_sheet]
            
            # 버튼 컨트롤
            button_cols = st.columns([1, 1, 5])  # 비율 조정
            with button_cols[0]:
                if st.button("Start", use_container_width=False):
                    st.session_state.playing = True
            with button_cols[1]:
                if st.button("Stop", use_container_width=False):
                    st.session_state.playing = False
                    pygame.mixer.music.stop()
                    pygame.mixer.quit()
                    self.cleanup()
            
            # 진행률 표시 (버튼과 텍스트 사이)
            st.write("")  # 작은 간격 추가
            total_rows = sheet.max_row - 1  # 헤더 제외
            current_progress = (st.session_state.current_row - 2) / total_rows if total_rows > 0 else 0
            
            # 슬라이더바와 진행 표시
            st.markdown("""
                <style>
                    .block-container {
                        padding-left: 1rem;
                        padding-right: 1rem;
                    }
                    .stProgress > div > div {
                        background-color: #1E1E1E;
                    }
                    .progress-container {
                        position: relative;
                        width: 720px;
                        padding-left: 0;
                        padding-right: 0;
                        margin-left: 0;
                        margin-right: 0;
                        margin-bottom: 25px;
                    }
                    .progress-number {
                        position: absolute;
                        left: 0;
                        right: 0;
                        top: -10px;
                        text-align: center;
                    }
                    .progress-start {
                        position: absolute;
                        left: 0;
                        bottom: -25px;
                        color: #FFFFFF;
                    }
                    .progress-end {
                        position: absolute;
                        right: 0;
                        bottom: -25px;
                        color: #FFFFFF;
                    }
                    .current-number {
                        background-color: #2E2E2E;
                        border-radius: 15px;
                        padding: 2px 10px;
                        color: #FFFFFF;
                        display: inline-block;
                        position: absolute;
                        left: calc(""" + str(current_progress * 100) + """% - 15px);
                        transform: translateX(-50%);
                    }
                    .stProgress {
                        max-width: 720px;
                        padding-left: 0;
                        padding-right: 0;
                        margin-left: 0;
                        margin-right: 0;
                    }
                    .element-container {
                        padding-left: 0;
                        padding-right: 0;
                    }
                </style>
                <div class="progress-container">
                    <div class="progress-number">
                        <span class="current-number">""" + str(st.session_state.current_row - 1) + """</span>
                    </div>
                    <div class="progress-start">1</div>
                    <div class="progress-end">""" + str(total_rows) + """</div>
                </div>
            """, unsafe_allow_html=True)
            
            st.progress(current_progress)
            st.write("")  # 작은 간격 추가
            
            # 텍스트 컨테이너 생성
            self.korean_container = st.empty()
            self.english_container = st.empty()
            
            # 오디오 컨테이너는 숨김 처리
            with st.empty():
                self.audio_container = st.container()
            
            # 재생 상태 확인 및 실행
            if st.session_state.playing:
                self.play_current_item(sheet)

    def play_current_item(self, sheet):
        if st.session_state.current_row <= sheet.max_row:
            try:
                # 현재 행의 텍스트 가져오기
                kor_text = sheet.cell(row=st.session_state.current_row, column=2).value
                eng_text = sheet.cell(row=st.session_state.current_row, column=3).value
                
                # 한글 텍스트 표시
                self.korean_container.markdown(f"<p style='color: white; font-size: 16px;'>{kor_text if kor_text else ''}</p>", unsafe_allow_html=True)
                self.english_container.empty()
                
                # 한글 음성 생성 및 재생
                if kor_text and st.session_state.playing:
                    korean_audio_placeholder = st.empty()
                    with korean_audio_placeholder.container():
                        temp_file = f"temp_ko_{st.session_state.current_row}.wav"
                        # 속도를 +20% 증가
                        cmd = f'edge-tts --voice "ko-KR-SunHiNeural" --text "{kor_text}" --rate="+20%" --write-media {temp_file}'
                        subprocess.run(cmd, shell=True)
                        
                        with open(temp_file, "rb") as audio_file:
                            audio_bytes = audio_file.read()
                            b64 = base64.b64encode(audio_bytes).decode()
                            st.markdown(
                                f"""
                                <audio autoplay="true" onended="this.remove();">
                                    <source src="data:audio/wav;base64,{b64}" type="audio/wav">
                                </audio>
                                """,
                                unsafe_allow_html=True
                            )
                        
                        os.remove(temp_file)
                        
                        # 재생 시간도 20% 감소
                        korean_chars = sum(1 for c in kor_text if ord('가') <= ord(c) <= ord('힣'))
                        other_chars = len(kor_text) - korean_chars
                        duration = ((korean_chars * 0.2) + (other_chars * 0.1) + 0.3) * 0.8
                        time.sleep(duration)
                    
                    korean_audio_placeholder.empty()
                
                if not st.session_state.playing:
                    return
                
                # 영어 텍스트 표시
                self.english_container.markdown(f"<p style='color: white; font-size: 16px;'>{eng_text if eng_text else ''}</p>", unsafe_allow_html=True)
                
                # 영어 음성 생성 및 재생
                if eng_text and st.session_state.playing:
                    english_audio_placeholder = st.empty()
                    with english_audio_placeholder.container():
                        temp_file = f"temp_en_{st.session_state.current_row}.wav"
                        # 속도를 +20% 증가
                        cmd = f'edge-tts --voice "en-US-JennyNeural" --text "{eng_text}" --rate="+20%" --write-media {temp_file}'
                        subprocess.run(cmd, shell=True)
                        
                        with open(temp_file, "rb") as audio_file:
                            audio_bytes = audio_file.read()
                            b64 = base64.b64encode(audio_bytes).decode()
                            st.markdown(
                                f"""
                                <audio autoplay="true" onended="this.remove();">
                                    <source src="data:audio/wav;base64,{b64}" type="audio/wav">
                                </audio>
                                """,
                                unsafe_allow_html=True
                            )
                        
                        os.remove(temp_file)
                        
                        # 재생 시간도 20% 감소
                        duration = ((len(eng_text) * 0.1) + 0.3) * 0.8
                        time.sleep(duration)
                    
                    english_audio_placeholder.empty()
                
                if st.session_state.playing:
                    st.session_state.current_row += 1
                    if st.session_state.current_row > sheet.max_row:
                        if st.session_state.repeat:
                            st.session_state.current_row = 2
                        else:
                            st.session_state.playing = False
                            return
                    
                    time.sleep(0.05)
                    st.rerun()
                
            except Exception as e:
                st.error(f"재생 오류: {e}")
                st.session_state.playing = False

    async def generate_and_play_speech(self, text, language):
        try:
            voice = "ko-KR-SunHiNeural" if language == "ko" else "en-US-JennyNeural"
            communicate = edge_tts.Communicate(text, voice)
            
            # 음성 생성
            audio_data = bytearray()
            async for chunk in communicate.stream():
                if chunk["type"] == "audio":
                    audio_data.extend(chunk["data"])
            
            # 음성 재생을 위한 HTML 생성
            b64_audio = base64.b64encode(audio_data).decode()
            audio_html = f"""
                <script>
                    var audio = new Audio("data:audio/mp3;base64,{b64_audio}");
                    audio.play().then(() => {{
                        audio.addEventListener('ended', () => {{
                            setTimeout(() => window.streamlit.scriptRunOnce(), 100);
                        }});
                    }});
                </script>
            """
            st.markdown(audio_html, unsafe_allow_html=True)
            
            # 음성 길이에 따른 대기
            await asyncio.sleep(len(text) * 0.1)
            
        except Exception as e:
            st.error(f"음성 생성 오류: {e}")

    def cleanup(self):
        try:
            for file in os.listdir(self.temp_dir):
                file_path = os.path.join(self.temp_dir, file)
                try:
                    os.remove(file_path)
                except:
                    pass
            os.rmdir(self.temp_dir)
        except:
            pass

def main():
    try:
        app = KorEngPlayer()
    except Exception as e:
        st.error(f"실행 중 오류 발생: {e}")

if __name__ == "__main__":
    main()
